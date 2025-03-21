from flask import render_template, request, redirect, url_for, session
from app import app
from app.forms import *
import os
from mysql.connector import Error
import mysql.connector
import openpyxl
import subprocess
from openpyxl.styles import NamedStyle
from datetime import datetime
import requests
import shutil


class Excel:
    def __init__(self):
        pass
        
    def Comp_Excel(self, nom_data, prenom_data, usine_data, client_data, contact_data, 
                  telephone_contact_data, email_contact, debut_data, fin_data, 
                  matricule_data, charge_data, adresse_data, mail_charge_data, 
                  telephone_charge_data, affaire_data, adresse_usine_data, 
                  mission_data, immatriculation_data, zone_data, excel_path=None):
        """Rempli le excel et le convertis en PDF"""
        # Utiliser le chemin Excel fourni ou utiliser celui par défaut
        self.excel_path = excel_path or os.path.abspath("app/formulaire.xlsx")
        
        # Vérifier que le fichier Excel existe
        if not os.path.isfile(self.excel_path):
            raise FileNotFoundError(f"Le fichier Excel {self.excel_path} n'existe pas")
        
        # Ouvrir le workbook spécifié
        self.workbook = openpyxl.load_workbook(self.excel_path)
        
        # Remplir les données
        self.nom(nom_data, prenom_data)
        #self.prenom(prenom_data)
        self.usine(usine_data, adresse_usine_data)
        self.affaire(affaire_data)
        self.client(client_data, contact_data, telephone_contact_data, email_contact)
        self.date_debut(debut_data)
        self.date_fin(fin_data)
        self.matricule(matricule_data)
        self.charge(charge_data)
        self.adresse(adresse_data)
        self.mission(mission_data)
        self.info_charge(telephone_charge_data, mail_charge_data)
        self.immatriculation(immatriculation_data)
        self.zone(zone_data)
        self.date()
        
        # Sauvegarder le workbook
        self.workbook.save(self.excel_path)
        
    def PDF(self, excel_path, odm_pdf):
        # Vérifier que le fichier Excel existe
        if not os.path.isfile(excel_path):
            print(f"Erreur: Le fichier Excel {excel_path} n'existe pas")
            return None
            
        # Créer le dossier de sortie s'il n'existe pas
        os.makedirs(odm_pdf, exist_ok=True)
        
        # Afficher les chemins pour le débogage
        print(f"Chemin Excel: {excel_path}")
        print(f"Dossier PDF: {odm_pdf}")
        
        # Conversion avec LibreOffice
        command = [
            "libreoffice", "--headless", "--convert-to", "pdf:writer_pdf_Export",
            excel_path, "--outdir", odm_pdf
        ]
        
        try:
            result = subprocess.run(
                command,
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )
            print("Conversion réussie. Sortie:", result.stdout)
            
            # Récupération du chemin du PDF généré
            pdf_filename = os.path.splitext(os.path.basename(excel_path))[0] + ".pdf"
            pdf_path = os.path.join(odm_pdf, pdf_filename)
            
            # Vérifier que le PDF a bien été créé
            if not os.path.isfile(pdf_path):
                print(f"Erreur: Le PDF {pdf_path} n'a pas été généré")
                return None
                
            return pdf_path
            
        except subprocess.CalledProcessError as e:
            print("Échec de la conversion. Erreur:", e.stderr)
            return None


    def date(self):
        # Charger le fichier Excel
        file_path = os.path.abspath("app/formulaire.xlsx")
        excel = openpyxl.load_workbook(file_path)

        # Obtenir la date actuelle
        date_actuelle = datetime.now()

        # Formater la date en "jour/mois/année"
        date_formatee = date_actuelle.strftime("%d/%m/%Y")

        # Sélectionner la feuille active
        sheet = excel.active

        sheet["AA6"] = date_formatee
        sheet["AB47"] = date_formatee
        sheet["H2"] = f'indice 1          date création : {date_formatee}'

        # Sauvegarder les modifications dans le fichier
        excel.save(file_path)

           

    def nom(self, nom, prenom):
         # Charger le fichier Excel
        file_path = os.path.abspath("app/formulaire.xlsx")
        print(file_path)
        excel = openpyxl.load_workbook(file_path)

        # Sélectionner la feuille active
        sheet = excel.active

        sheet["H6"] = f'{nom} {prenom}'

        # Sauvegarder les modifications dans le fichier
        excel.save(file_path)

    def usine(self, usine, adresse_usine):
         # Charger le fichier Excel
        file_path = os.path.abspath("app/formulaire.xlsx")
        excel = openpyxl.load_workbook(file_path)

        # Sélectionner la feuille active
        sheet = excel.active

        sheet["E12"] = usine
        sheet["A13"] = adresse_usine

        # Sauvegarder les modifications dans le fichier
        excel.save(file_path)

    def client(self, client, contact, telephone_contact,email_contact):
         # Charger le fichier Excel
        file_path = os.path.abspath("app/formulaire.xlsx")
        excel = openpyxl.load_workbook(file_path)

        # Sélectionner la feuille active
        sheet = excel.active

        sheet["D10"] = client
        sheet["X12"] = contact
        sheet["S14"] = telephone_contact
        sheet["AA14"] = email_contact
        
        # Sauvegarder les modifications dans le fichier
        excel.save(file_path)

    def date_debut(self, debut):
        # Charger le fichier Excel
        file_path = os.path.abspath("app/formulaire.xlsx")
        excel = openpyxl.load_workbook(file_path)

        # Sélectionner la feuille active
        sheet = excel.active

        # Écriture de la data dans la cellule
        sheet["F17"] = debut

        # Sauvegarder les modifications dans le fichier
        excel.save(file_path)

    def date_fin(self, fin):
        # Charger le fichier Excel
        file_path = os.path.abspath("app/formulaire.xlsx")
        excel = openpyxl.load_workbook(file_path)

        # Sélectionner la feuille active
        sheet = excel.active

        # Écriture de la data dans la cellule
        sheet["Y17"] = fin

        # Sauvegarder les modifications dans le fichier
        excel.save(file_path)


    def matricule(self, data_matricule):
        # Charger le fichier Excel
        file_path = os.path.abspath("app/formulaire.xlsx")
        excel = openpyxl.load_workbook(file_path)

        # Sélectionner la feuille active
        sheet = excel.active

        # Écriture de la data dans la cellule
        sheet["H5"] = data_matricule

        # Sauvegarder les modifications dans le fichier
        excel.save(file_path)

    def charge(self, data_charge):
        # Charger le fichier Excel
        file_path = os.path.abspath("app/formulaire.xlsx")
        excel = openpyxl.load_workbook(file_path)

        # Sélectionner la feuille active
        sheet = excel.active

        # Écriture de la data dans la cellule
        sheet["W10"] = data_charge

        # Sauvegarder les modifications dans le fichier
        excel.save(file_path)

    def info_charge(self, telephone, mail):
        # Charger le fichier Excel
        file_path = os.path.abspath("app/formulaire.xlsx")
        excel = openpyxl.load_workbook(file_path)

        # Sélectionner la feuille active
        sheet = excel.active

        # Écriture de la data dans la cellule
        sheet["AD11"] = telephone
        sheet["W11"] = mail

        # Sauvegarder les modifications dans le fichier
        excel.save(file_path)

    def adresse(self, data_adresse):
        # Charger le fichier Excel
        file_path = os.path.abspath("app/formulaire.xlsx")
        excel = openpyxl.load_workbook(file_path)

        # Sélectionner la feuille active
        sheet = excel.active

        # Écriture de la data dans la cellule
        sheet["H7"] = data_adresse

        # Sauvegarder les modifications dans le fichier
        excel.save(file_path)

    def affaire(self, data_affaire):
        # Charger le fichier Excel
        file_path = os.path.abspath("app/formulaire.xlsx")
        excel = openpyxl.load_workbook(file_path)

        # Sélectionner la feuille active
        sheet = excel.active

        # Écriture de la data dans la cellule
        sheet["AC4"] = data_affaire
        sheet["D11"] = data_affaire

        # Sauvegarder les modifications dans le fichier
        excel.save(file_path)

    def mission(self, data_mission):
        # Charger le fichier Excel
        file_path = os.path.abspath("app/formulaire.xlsx")
        excel = openpyxl.load_workbook(file_path)

        # Sélectionner la feuille active
        sheet = excel.active

        # Écriture de la data dans la cellule
        sheet["I18"] = data_mission

        # Sauvegarder les modifications dans le fichier
        excel.save(file_path)

    def immatriculation(self, data_immatriculation):
        # Charger le fichier Excel
        file_path = os.path.abspath("app/formulaire.xlsx")
        excel = openpyxl.load_workbook(file_path)

        # Sélectionner la feuille active
        sheet = excel.active

        # Écriture de la data dans la cellule
        sheet["N34"] = data_immatriculation

        # Sauvegarder les modifications dans le fichier
        excel.save(file_path)

    def zone(self, data_zone):
        # Charger le fichier Excel
        file_path = os.path.abspath("app/formulaire.xlsx")
        excel = openpyxl.load_workbook(file_path)

        # Sélectionner la feuille active
        sheet = excel.active

        # Écriture de la data dans la cellule
        sheet["P20"] = data_zone

        # Sauvegarder les modifications dans le fichier
        excel.save(file_path)
